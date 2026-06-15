# Importação Django (acabaram as minhas piadas - Desculpa)
from datetime import datetime
from io import BytesIO

from django.shortcuts import render
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importações do código
from ..models import Estoque, BensConsumo

# ===================================================
# CAMPOS DESTINADOS PARA GERENCIAR PÁGINAS INICIAIS
# ===================================================


''' Homepage '''
# Página Principail
@login_required
def homepage(request):

    query = request.GET.get('q', '').strip()
    grupo = request.GET.get('grupo', '').strip()

    itens = Estoque.objects.select_related('item_shock', 'locate').all()

    # Busca
    if query:
        itens = itens.filter(
            Q(item_shock__efisco__icontains=query) |
            Q(mark__icontains=query) |
            Q(description_manual__icontains=query)
        )

    # Filtro por grupo
    if grupo:
        itens = itens.filter(item_shock__grupo_consumo=grupo)

    grupos = BensConsumo._meta.get_field("grupo_consumo").choices

    # itens essenciais
    item_essential = itens.filter(essential=True)
    
    # estoque baixo 
    estoque_baixo = [item for item in itens if item.low_stock]
    alerta_vencimento = [item for item in itens if item.alerta_vencimento]

    # ordenar alerta_vencimento pela data mais próxima / vencida no topo 
    alerta_vencimento = sorted(alerta_vencimento, key=lambda x: x.validity)

    # ordenar estoque_baixo pelos dias restantes, do maior para o menor
    estoque_baixo = sorted(estoque_baixo, key=lambda x: x.duration, reverse=True)

    context = {
        'itens': itens,
        'total_itens': itens.count(),
        'query': query,
        'grupos': grupos,
        'grupo_selected': grupo,
        'item_essential': item_essential,
        'estoque_baixo': estoque_baixo,
        'alerta_vencimento': alerta_vencimento,
    }

    return render(request, 'dimms/homepage.html', context)


''' Exportar Planilha '''
@login_required
def exportar_planilha(request):
    itens = (
        Estoque.objects
        .select_related('item_shock', 'locate__setor_sala')
        .order_by('item_shock__grupo_consumo', 'description_manual')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Estoque'

    # ── estilos ───────────────────────────────────────────────────────────────
    navy        = PatternFill('solid', fgColor='1E2D42')
    stripe      = PatternFill('solid', fgColor='F0F4FB')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    cell_font   = Font(name='Calibri', size=10)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='C8CFE0')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        ('Essencial',          12),
        ('Almoxarifado',       22),
        ('Grupo',              20),
        ('E-Fisco',            14),
        ('Descrição',          40),
        ('Unidade',            14),
        ('Endereço',           22),
        ('Marca',              18),
        ('Cons. mes. hist.',   16),
        ('Consumo mensal',     16),
        ('Custo unit.',        14),
        ('Duração',            16),
        ('Validade',           14),
        ('Qtd. Estoque',       14),
    ]

    # cabeçalho
    for col, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill      = navy
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # dados
    for row_idx, item in enumerate(itens, start=2):
        shade = (row_idx % 2 == 0)
        bg    = stripe if shade else PatternFill('solid', fgColor='FFFFFF')

        locate     = item.locate
        almox      = item.item_shock.get_almoxarifado_display() if item.item_shock.almoxarifado else '—'
        endereco   = str(locate) if locate else '—'
        grupo      = item.item_shock.get_grupo_consumo_display() if item.item_shock.grupo_consumo else '—'
        unidade    = item.item_shock.get_medida_display()        if item.item_shock.medida        else '—'
        validade   = item.validity.strftime('%d/%m/%Y')          if item.validity                 else '—'
        duracao    = str(item.duration)                          if item.duration                 else '—'

        values = [
            'Sim' if item.essential else 'Não',
            almox,
            grupo,
            item.item_shock.efisco or '—',
            item.description_manual or '—',
            unidade,
            endereco,
            item.mark or '—',
            item.monthly_consumption if item.monthly_consumption is not None else '—',
            item.monthly_consumption if item.monthly_consumption is not None else '—',
            '',        # Custo unit. — sem campo no modelo
            duracao,
            validade,
            item.amount_shock,
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = cell_font
            cell.fill      = bg
            cell.border    = border
            cell.alignment = center if col in (1, 6, 9, 10, 11, 13, 14) else left

    # ── aba de metadados ──────────────────────────────────────────────────────
    ws_meta = wb.create_sheet('Info')
    ws_meta['A1'] = 'Gerado em'
    ws_meta['B1'] = datetime.now().strftime('%d/%m/%Y às %H:%M')
    ws_meta['A2'] = 'Total de itens'
    ws_meta['B2'] = itens.count()
    ws_meta['A1'].font = Font(bold=True)
    ws_meta['A2'].font = Font(bold=True)
    ws_meta.column_dimensions['A'].width = 18
    ws_meta.column_dimensions['B'].width = 24

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'estoque_dimms_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


''' Páginas de Alertas '''
# Estoque Baixo
@login_required
def low_stock(request):
    query   = request.GET.get('q', '').strip()
    grupo_f = request.GET.get('grupo', '').strip()
    essencial_f = request.GET.get('essencial', '').strip()

    itens = (
        Estoque.objects
        .select_related('item_shock', 'locate')
        .filter(monthly_consumption__isnull=False)
        .order_by('description_manual')
    )

    if query:
        itens = itens.filter(
            Q(description_manual__icontains=query) |
            Q(item_shock__efisco__icontains=query) |
            Q(mark__icontains=query)
        )

    if grupo_f:
        itens = itens.filter(item_shock__grupo_consumo=grupo_f)

    if essencial_f == '1':
        itens = itens.filter(essential=True)
    elif essencial_f == '0':
        itens = itens.filter(essential=False)

    estoque_baixo = [i for i in itens if i.low_stock]

    total_baixo   = len(estoque_baixo)
    tb_essenciais = sum(1 for i in estoque_baixo if i.essential)
    tb_vencimento = sum(1 for i in estoque_baixo if i.alerta_vencimento)

    from dimms.utils import GrupoConsumo
    grupos = GrupoConsumo.choices

    context = {
        'estoque_baixo': estoque_baixo,
        'total_baixo': total_baixo,
        'tb_essenciais': tb_essenciais,
        'tb_vencimento': tb_vencimento,
        'query': query,
        'grupo_f': grupo_f,
        'essencial_f': essencial_f,
        'grupos': grupos,
    }

    return render(request, 'dimms/alerts/low_stock.html', context)

# Itens Essenciais 
@login_required 
def essential(request):
    query      = request.GET.get('q', '').strip()
    grupo_f    = request.GET.get('grupo', '').strip()
    alerta_f   = request.GET.get('alerta', '').strip()

    itens = Estoque.objects.select_related('item_shock', 'locate').filter(essential=True)

    if query:
        itens = itens.filter(
            Q(description_manual__icontains=query) |
            Q(item_shock__efisco__icontains=query) |
            Q(mark__icontains=query)
        )

    if grupo_f:
        itens = itens.filter(item_shock__grupo_consumo=grupo_f)

    # contagens para KPIs (antes do filtro de alerta)
    total_itens   = itens.count()
    estoque_baixo = sum(1 for i in itens if i.low_stock)
    vence_em_breve = sum(1 for i in itens if i.alerta_vencimento)

    # filtro de alerta aplicado após KPIs
    if alerta_f == 'baixo':
        itens = [i for i in itens if i.low_stock]
    elif alerta_f == 'vencimento':
        itens = [i for i in itens if i.alerta_vencimento]
    else:
        itens = list(itens)

    from dimms.utils import GrupoConsumo
    grupos = GrupoConsumo.choices

    context = {
        'itens': itens,
        'total_itens': total_itens,
        'estoque_baixo': estoque_baixo,
        'vence_em_breve': vence_em_breve,
        'query': query,
        'grupo_f': grupo_f,
        'alerta_f': alerta_f,
        'grupos': grupos,
    }

    return render(request, 'dimms/alerts/essential.html', context)

# Itens perto do vencimento
@login_required    
def expiration_alert(request):
    from datetime import date
    query       = request.GET.get('q', '').strip()
    grupo_f     = request.GET.get('grupo', '').strip()
    status_f    = request.GET.get('status', '').strip()  # vencido | a_vencer
    essencial_f = request.GET.get('essencial', '').strip()

    itens = (
        Estoque.objects
        .select_related('item_shock', 'locate')
        .filter(validity__isnull=False)
        .order_by('validity', 'description_manual')
    )

    if query:
        itens = itens.filter(
            Q(description_manual__icontains=query) |
            Q(item_shock__efisco__icontains=query) |
            Q(mark__icontains=query)
        )

    if grupo_f:
        itens = itens.filter(item_shock__grupo_consumo=grupo_f)

    if essencial_f == '1':
        itens = itens.filter(essential=True)
    elif essencial_f == '0':
        itens = itens.filter(essential=False)

    hoje = date.today()
    alerta_vencimento = [i for i in itens if i.alerta_vencimento]

    total_alerta  = len(alerta_vencimento)
    total_vencidos = sum(1 for i in alerta_vencimento if i.validity < hoje)
    total_a_vencer = sum(1 for i in alerta_vencimento if i.validity >= hoje)
    total_essenciais = sum(1 for i in alerta_vencimento if i.essential)

    if status_f == 'vencido':
        alerta_vencimento = [i for i in alerta_vencimento if i.validity < hoje]
    elif status_f == 'a_vencer':
        alerta_vencimento = [i for i in alerta_vencimento if i.validity >= hoje]

    from dimms.utils import GrupoConsumo
    grupos = GrupoConsumo.choices

    context = {
        'alerta_vencimento': alerta_vencimento,
        'total_alerta': total_alerta,
        'total_vencidos': total_vencidos,
        'total_a_vencer': total_a_vencer,
        'total_essenciais': total_essenciais,
        'query': query,
        'grupo_f': grupo_f,
        'status_f': status_f,
        'essencial_f': essencial_f,
        'grupos': grupos,
        'hoje': hoje,
    }

    return render(request, 'dimms/alerts/expiration_alert.html', context)