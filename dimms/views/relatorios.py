# Importações do Django
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas

# Importações do código
from ..models import Estoque, SolicitacaoItens
from ..utils import GrupoConsumo
from .saldo_ativo import _pdf_header, _fmt_dt
from dempam.models import CircunscricaoPredio, InfoUA

MESES_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro',
}

# ====================================
# RELATÓRIOS DO DIMMS (BENS DE CONSUMO)
# ====================================


def _fmt_qtd(valor):
    if valor is None:
        return '—'
    texto = f'{valor:.2f}'.rstrip('0').rstrip('.')
    return texto or '0'


def _truncar(c, texto, largura_max, fonte='Helvetica', tamanho=8):
    if c.stringWidth(texto, fonte, tamanho) <= largura_max:
        return texto
    while texto and c.stringWidth(texto + '…', fonte, tamanho) > largura_max:
        texto = texto[:-1]
    return (texto + '…') if texto else '…'


''' Relatórios — página inicial com os relatórios disponíveis '''
@login_required
def relatorios(request):
    return render(request, 'dimms/relatorios/relatorios.html')


ICONES_GRUPO = {
    GrupoConsumo.alimento: '🍽️',
    GrupoConsumo.confeccao: '🧵',
    GrupoConsumo.copa_cozinha: '☕',
    GrupoConsumo.domissanitario: '🧼',
    GrupoConsumo.eletrica: '🔌',
    GrupoConsumo.epi: '🦺',
    GrupoConsumo.hidrosanitario: '🚿',
    GrupoConsumo.informatica: '🖥️',
    GrupoConsumo.limpeza: '🧹',
    GrupoConsumo.manutencao: '🛠️',
    GrupoConsumo.marcenaria: '🪚',
    GrupoConsumo.pintura: '🎨',
    GrupoConsumo.papeis_expediente: '📄',
    GrupoConsumo.papeis_limpeza: '🧻',
    GrupoConsumo.refrigeracao: '❄️',
    GrupoConsumo.toner: '🖨️',
}


''' Ajuste de Inventário — seleção de grupo '''
@login_required
def ajuste_inventario(request):
    contagem = dict(
        Estoque.objects
        .values_list('item_shock__grupo_consumo')
        .annotate(total=Count('id'))
        .values_list('item_shock__grupo_consumo', 'total')
    )
    grupos = [
        {
            'value': value,
            'label': label,
            'total': contagem.get(value, 0),
            'icone': ICONES_GRUPO.get(value, '📦'),
        }
        for value, label in GrupoConsumo.choices
    ]
    return render(request, 'dimms/relatorios/ajuste_inventario.html', {
        'grupos': grupos,
        'total_itens': Estoque.objects.count(),
    })


''' Ajuste de Inventário — PDF de contagem física por grupo(s) '''
@login_required
def ajuste_inventario_pdf(request):
    grupos_sel = [g for g in request.GET.getlist('grupo') if g]

    itens = Estoque.objects.select_related('item_shock').order_by('description_manual')
    if grupos_sel:
        itens = itens.filter(item_shock__grupo_consumo__in=grupos_sel)

    labels_grupo = dict(GrupoConsumo.choices)
    if not grupos_sel:
        grupo_label = 'Todos os grupos'
        arquivo_slug = 'todos'
    elif len(grupos_sel) == 1:
        grupo_label = labels_grupo.get(grupos_sel[0], grupos_sel[0])
        arquivo_slug = grupos_sel[0]
    else:
        nomes = [labels_grupo.get(g, g) for g in grupos_sel]
        if len(nomes) > 4:
            grupo_label = f'{", ".join(nomes[:4])} e mais {len(nomes) - 4}'
        else:
            grupo_label = ', '.join(nomes)
        arquivo_slug = f'{len(grupos_sel)}-grupos'

    resp = HttpResponse(content_type='application/pdf')
    filename = f'ajuste_inventario_{arquivo_slug}.pdf'.lower()
    resp['Content-Disposition'] = f'inline; filename="{filename}"'

    w, h_page = landscape(A4)
    c = rl_canvas.Canvas(resp, pagesize=(w, h_page))

    col_efisco = 12 * mm
    col_nome = col_efisco + 30 * mm
    col_qtd = col_nome + 128 * mm
    col_ajuste = col_qtd + 30 * mm
    row_h = 7 * mm

    rotulo_subtitulo = 'Grupos' if len(grupos_sel) > 1 else 'Grupo'

    def cabecalho_pagina():
        y = _pdf_header(
            c, w, h_page,
            titulo='Ajuste de Inventário — Bens de Consumo',
            subtitulo=f'{rotulo_subtitulo}: {grupo_label}',
            data_str=f'Gerado em {_fmt_dt(timezone.now())}',
        )
        return cabecalho_tabela(y)

    def cabecalho_tabela(y):
        c.setFillColor(colors.HexColor('#1e2d42'))
        c.rect(12 * mm, y - row_h, w - 24 * mm, row_h, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 8)
        c.drawString(col_efisco + 3 * mm, y - 4.8 * mm, 'E-FISCO')
        c.drawString(col_nome + 3 * mm, y - 4.8 * mm, 'DESCRIÇÃO')
        c.drawRightString(col_qtd + 24 * mm, y - 4.8 * mm, 'QTD. ESTOQUE')
        c.drawString(col_ajuste + 3 * mm, y - 4.8 * mm, 'AJUSTE (PREENCHER)')
        return y - row_h

    y = cabecalho_pagina()
    c.setFont('Helvetica', 8)

    total = 0
    for idx, item in enumerate(itens):
        if y < 25 * mm:
            c.showPage()
            y = cabecalho_pagina()
            c.setFont('Helvetica', 8)

        if idx % 2 == 0:
            c.setFillColor(colors.HexColor('#f7f8fc'))
            c.rect(12 * mm, y - row_h, w - 24 * mm, row_h, fill=1, stroke=0)

        c.setFillColor(colors.HexColor('#1a2035'))
        c.setFont('Helvetica', 8)
        c.drawString(col_efisco + 3 * mm, y - 4.8 * mm, item.item_shock.efisco if item.item_shock else '—')
        nome = _truncar(c, item.description_manual or '—', col_qtd - col_nome - 6 * mm)
        c.drawString(col_nome + 3 * mm, y - 4.8 * mm, nome)
        c.drawRightString(col_qtd + 24 * mm, y - 4.8 * mm, _fmt_qtd(item.amount_shock))

        c.setStrokeColor(colors.HexColor('#c8cfe0'))
        c.setLineWidth(0.3)
        c.line(12 * mm, y - row_h, w - 12 * mm, y - row_h)
        c.line(col_ajuste, y, col_ajuste, y - row_h)

        y -= row_h
        total += 1

    c.setFillColor(colors.HexColor('#7380a0'))
    c.setFont('Helvetica', 7.5)
    c.drawString(12 * mm, 12 * mm, f'Total de itens: {total}')

    c.save()
    return resp


''' Custo por Região — soma o preço médio × quantidade pedida de todas as UAs de uma circunscrição '''
@login_required
def custo_por_regiao(request):
    circunscricoes = CircunscricaoPredio.objects.all().order_by('local')

    circ_id = request.GET.get('circunscricao', '').strip()
    data_de = request.GET.get('data_de', '').strip()
    data_ate = request.GET.get('data_ate', '').strip()

    circunscricao_selecionada = None
    linhas = []
    total_geral = Decimal('0')
    total_qtd_geral = Decimal('0')
    qtd_solicitacoes_geral = 0
    itens_sem_preco = 0
    bens_sem_preco = set()

    if circ_id:
        circunscricao_selecionada = CircunscricaoPredio.objects.filter(pk=circ_id).first()

    if circunscricao_selecionada:
        uas = InfoUA.objects.filter(circunscricao_predio=circunscricao_selecionada)

        itens = (
            SolicitacaoItens.objects
            .filter(
                request_defendant__ua_order__in=uas,
                item_order__isnull=False,
                amount_order__isnull=False,
            )
            .select_related('item_order__item_shock', 'request_defendant__ua_order')
        )

        if data_de:
            itens = itens.filter(request_defendant__data_order__date__gte=data_de)
        if data_ate:
            itens = itens.filter(request_defendant__data_order__date__lte=data_ate)

        por_ua = {}
        for it in itens:
            ua = it.request_defendant.ua_order
            if ua is None:
                continue

            linha = por_ua.setdefault(ua.pk, {
                'ua': ua,
                'total_qtd': Decimal('0'),
                'total_custo': Decimal('0'),
                'solicitacoes': set(),
            })
            linha['total_qtd'] += it.amount_order
            linha['solicitacoes'].add(it.request_defendant_id)

            preco = it.item_order.item_shock.preco_medio
            if preco is not None:
                linha['total_custo'] += it.amount_order * preco
            else:
                itens_sem_preco += 1
                bens_sem_preco.add(it.item_order.item_shock_id)

        linhas = sorted(por_ua.values(), key=lambda l: l['total_custo'], reverse=True)
        for l in linhas:
            l['qtd_solicitacoes'] = len(l['solicitacoes'])

        total_geral = sum((l['total_custo'] for l in linhas), Decimal('0'))
        total_qtd_geral = sum((l['total_qtd'] for l in linhas), Decimal('0'))
        qtd_solicitacoes_geral = len(set().union(*[l['solicitacoes'] for l in linhas])) if linhas else 0

    context = {
        'circunscricoes': circunscricoes,
        'circunscricao_selecionada': circunscricao_selecionada,
        'data_de': data_de,
        'data_ate': data_ate,
        'linhas': linhas,
        'total_geral': total_geral,
        'total_qtd_geral': total_qtd_geral,
        'qtd_solicitacoes_geral': qtd_solicitacoes_geral,
        'itens_sem_preco': itens_sem_preco,
        'qtd_bens_sem_preco': len(bens_sem_preco),
    }
    return render(request, 'dimms/relatorios/custo_por_regiao.html', context)


# Agrega SolicitacaoItens por item (E-Fisco) e mês do pedido — usada tanto pela
# tela do relatório (paginada) quanto pela exportação XLSX (linha completa, sem paginar)
def _montar_gastos_por_item(request):
    data_de = request.GET.get('data_de', '').strip()
    data_ate = request.GET.get('data_ate', '').strip()
    grupo_sel = request.GET.get('grupo', '').strip()
    busca = request.GET.get('busca', '').strip()

    itens = (
        SolicitacaoItens.objects
        .filter(item_order__isnull=False, amount_order__isnull=False, request_defendant__isnull=False)
        .select_related('item_order__item_shock', 'request_defendant')
    )
    if data_de:
        itens = itens.filter(request_defendant__data_order__date__gte=data_de)
    if data_ate:
        itens = itens.filter(request_defendant__data_order__date__lte=data_ate)
    if grupo_sel:
        itens = itens.filter(item_order__item_shock__grupo_consumo=grupo_sel)
    if busca:
        itens = itens.filter(
            Q(item_order__item_shock__efisco__icontains=busca) |
            Q(item_order__item_shock__descricao_efisco__icontains=busca)
        )

    por_item_mes = {}
    itens_sem_preco = set()
    for it in itens:
        efisco_obj = it.item_order.item_shock
        data_pedido = it.request_defendant.data_order
        if data_pedido is None:
            continue
        ano, mes = data_pedido.year, data_pedido.month

        linha = por_item_mes.setdefault((efisco_obj.pk, ano, mes), {
            'item': efisco_obj,
            'ano': ano,
            'mes': mes,
            'total_qtd': Decimal('0'),
            'total_gasto': Decimal('0'),
        })
        linha['total_qtd'] += it.amount_order

        preco = efisco_obj.preco_medio
        if preco is not None:
            linha['total_gasto'] += it.amount_order * preco
        else:
            itens_sem_preco.add(efisco_obj.pk)

    linhas = sorted(por_item_mes.values(), key=lambda l: (l['ano'], l['mes'], l['total_gasto']), reverse=True)

    total_geral = sum((l['total_gasto'] for l in linhas), Decimal('0'))
    total_qtd_geral = sum((l['total_qtd'] for l in linhas), Decimal('0'))

    return {
        'linhas': linhas,
        'data_de': data_de,
        'data_ate': data_ate,
        'grupo_sel': grupo_sel,
        'busca': busca,
        'total_geral': total_geral,
        'total_qtd_geral': total_qtd_geral,
        'qtd_itens_sem_preco': len(itens_sem_preco),
    }


GASTOS_POR_ITEM_POR_PAGINA = 30


''' Gastos por Item — quantidade saída × preço médio, agrupado por item e por mês '''
@login_required
def gastos_por_item(request):
    dados = _montar_gastos_por_item(request)

    pagina = Paginator(dados['linhas'], GASTOS_POR_ITEM_POR_PAGINA).get_page(request.GET.get('pagina'))

    return render(request, 'dimms/relatorios/gastos_por_item.html', {
        **dados,
        'pagina': pagina,
        'grupos': GrupoConsumo.choices,
        'meses_pt': MESES_PT,
    })


''' Gastos por Item — exportação XLSX (linha completa, respeitando os filtros da tela) para uso externo (ex.: Power BI) '''
@login_required
def gastos_por_item_export_xlsx(request):
    dados = _montar_gastos_por_item(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gastos por Item'

    colunas = [
        ('E-Fisco', 16),
        ('Descrição', 60),
        ('Grupo', 22),
        ('Ano', 8),
        ('Mês', 14),
        ('Quantidade', 14),
        ('Preço Médio', 14),
        ('Gasto', 14),
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E2D42', end_color='1E2D42', fill_type='solid')

    for col_idx, (titulo, largura) in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=col_idx, value=titulo)
        cel.font = header_font
        cel.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = 'A2'

    labels_grupo = dict(GrupoConsumo.choices)
    for row_idx, linha in enumerate(dados['linhas'], start=2):
        item = linha['item']
        ws.cell(row=row_idx, column=1, value=item.efisco)
        ws.cell(row=row_idx, column=2, value=item.descricao_efisco)
        ws.cell(row=row_idx, column=3, value=labels_grupo.get(item.grupo_consumo, item.grupo_consumo))
        ws.cell(row=row_idx, column=4, value=linha['ano'])
        ws.cell(row=row_idx, column=5, value=MESES_PT[linha['mes']])
        ws.cell(row=row_idx, column=6, value=float(linha['total_qtd']))
        ws.cell(row=row_idx, column=7, value=float(item.preco_medio) if item.preco_medio is not None else None)
        ws.cell(row=row_idx, column=8, value=float(linha['total_gasto']))

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="gastos_por_item_sipat.xlsx"'
    wb.save(resp)
    return resp
