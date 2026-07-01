"""
Management command: importar_saldo_ativo_xlsx

Importa contratos, fornecedores e itens de saldo ativo a partir de uma
planilha XLSX no formato padrão da DIMMS.

Uso:
    python manage.py importar_saldo_ativo_xlsx [caminho.xlsx] [--sobrescrever]

O arquivo padrão é: docs/Saldo ativo de materiais 2026.xlsx
"""

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from dimms.models import (
    BensConsumo,
    Contrato,
    SaldoAtivo,
)
from dimms.models.artifacts import Supplier


# ─── Utilitários ────────────────────────────────────────────────────────────

def _normalize(text):
    """Remove barras, espaços extras e converte para minúsculas — usado para
    o matching fuzzy entre nomes do RESUMO e nomes das abas."""
    if not text:
        return ''
    return re.sub(r'[\s/]+', '', str(text)).lower()


def _to_decimal(value):
    if value is None:
        return None
    try:
        return Decimal(str(value)).quantize(Decimal('0.01'))
    except InvalidOperation:
        return None


def _parse_vigencia(text):
    """'03/12/2025 a 02/12/2026' → (date, date)"""
    if not text or not isinstance(text, str):
        return None, None
    m = re.search(r'(\d{2}/\d{2}/\d{4})\s+[aA]\s+(\d{2}/\d{2}/\d{4})', text.strip())
    if m:
        try:
            inicio = datetime.strptime(m.group(1), '%d/%m/%Y').date()
            fim = datetime.strptime(m.group(2), '%d/%m/%Y').date()
            return inicio, fim
        except ValueError:
            pass
    return None, None


def _cell(row, idx, default=None):
    """Lê uma célula de forma segura."""
    try:
        v = row[idx]
        if v is None:
            return default
        if isinstance(v, str):
            v = v.strip()
            return v if v else default
        return v
    except IndexError:
        return default


# ─── Leitura do RESUMO ──────────────────────────────────────────────────────

def _build_resumo_map(wb):
    """Retorna dict {nome_normalizado → dados_do_resumo}."""
    if 'RESUMO' not in wb.sheetnames:
        return {}

    ws = wb['RESUMO']
    rows = list(ws.iter_rows(values_only=True))
    result = {}

    for row in rows[2:]:           # linha 3 em diante (pula título + cabeçalho)
        if row[0] is None:         # linha de total ou vazia
            continue
        forma     = _cell(row, 0)
        fornecedor = _cell(row, 1)
        n_mppe    = _cell(row, 2)
        if not n_mppe:
            continue

        data = {
            'forma':         str(forma).upper() if forma else '',
            'fornecedor':    str(fornecedor) if fornecedor else '',
            'numero_processo': str(_cell(row, 3) or '').strip(),
            'numero_arp':      str(_cell(row, 4) or '').strip().lstrip('-').strip(),
            'categoria':       str(_cell(row, 5) or '').strip(),
            'vencimento':      row[6] if isinstance(row[6], datetime) else None,
            'aditivo_prazo':   str(_cell(row, 7) or '').strip(),
            'aditivo_preco':   str(_cell(row, 8) or '').strip(),
            'valor_contratado': _to_decimal(row[9]),
            'previsao_anual':   _to_decimal(row[13]),
        }
        key = _normalize(n_mppe)
        result[key] = data          # última entrada ganha (duplicatas raras)
        # Guarda também com n_mppe original para log
        data['_n_mppe'] = str(n_mppe).strip()

    return result


# ─── Leitura do cabeçalho de cada aba ──────────────────────────────────────

def _parse_header(rows):
    """Extrai metadados do cabeçalho padrão das abas de detalhe."""
    header = {}
    for row in rows[:20]:
        label = str(row[0] or '').strip().upper().rstrip(':')
        value = row[1]
        if 'FORNECEDOR' in label:
            header['fornecedor'] = str(value or '').strip()
        elif label == 'CNPJ':
            header['cnpj'] = str(value or '').strip()
        elif 'CONTATO' in label:
            header['contato'] = str(value or '').strip()
        elif 'E-MAIL' in label or 'EMAIL' in label:
            header['email'] = str(value or '').strip()
        elif 'N° CONTRATO' in label or 'N? CONTRATO' in label or 'CONTRATO' in label and 'N' in label:
            header['numero_contrato'] = str(value or '').strip()
        elif 'HOMOLOGA' in label:
            if isinstance(value, datetime):
                header['homologacao'] = value.date()
        elif 'VIG' in label:
            inicio, fim = _parse_vigencia(str(value or ''))
            if inicio:
                header['inicio_vigencia'] = inicio
                header['final_vigencia'] = fim
        elif label == 'N° SC' or label == 'N? SC':
            header['numero_sc'] = str(value or '').strip()
        elif 'COD' in label and 'LICITA' in label:
            raw = str(value or '').strip()
            header['cod_liquidacao'] = int(raw) if raw.isdigit() else None

    return header


# ─── Leitura dos itens de cada aba ─────────────────────────────────────────

def _parse_items(rows):
    """Encontra a linha de cabeçalho com 'E-FISCO' e extrai os itens abaixo."""
    header_idx = None
    for i, row in enumerate(rows):
        if any('E-FISCO' in str(c).upper() for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        return []

    items = []
    for row in rows[header_idx + 1:]:
        # Linha de item: primeira coluna é número (1, 2, 3...)
        if not isinstance(row[0], (int, float)):
            continue
        efisco_raw = _cell(row, 1)
        if efisco_raw is None:
            continue
        # E-Fisco vem como float (ex: 5832527.0) — converter p/ string inteira
        if isinstance(efisco_raw, float):
            efisco_str = str(int(efisco_raw))
        else:
            efisco_str = str(efisco_raw).strip().split('.')[0]

        items.append({
            'numero_item':  int(row[0]),
            'efisco':       efisco_str,
            'descricao':    str(_cell(row, 2) or '').replace('\n', ' ').strip(),
            'marca':        str(_cell(row, 3) or '').strip(),
            'quantidade':   int(row[5]) if isinstance(row[5], (int, float)) else None,
            'preco_unit':   _to_decimal(row[6]),
            'qtd_saldo':    int(row[8]) if isinstance(row[8], (int, float)) else None,
        })
    return items


# ─── Persistência ───────────────────────────────────────────────────────────

def _get_or_create_supplier(header, fornecedor_nome, sobrescrever):
    """Busca fornecedor por nome. Cria se não existir."""
    nome = (header.get('fornecedor') or fornecedor_nome or '').strip()
    if not nome:
        return None

    nome = nome[:100]   # respeita max_length

    supplier, created = Supplier.objects.get_or_create(
        supplier=nome,
        defaults={
            'cnpj_supplier':    (header.get('cnpj') or '')[:18],
            'name_responsible': (header.get('contato') or '')[:90],
            'email_supplier':   (header.get('email') or '')[:254],
            'contact_supplier': '',
            'cpf_responsible':  '',
        }
    )

    if not created and sobrescrever:
        if header.get('cnpj'):
            supplier.cnpj_supplier = (header.get('cnpj') or '')[:18]
        if header.get('contato'):
            supplier.name_responsible = (header.get('contato') or '')[:90]
        if header.get('email'):
            supplier.email_supplier = (header.get('email') or '')[:254]
        supplier.save(update_fields=['cnpj_supplier', 'name_responsible', 'email_supplier'])

    return supplier


def _get_or_create_contrato(n_mppe, header, resumo_data, supplier, sobrescrever):
    """Busca contrato pelo N° MPPE. Cria/atualiza com dados do resumo e cabeçalho."""
    contrato, created = Contrato.objects.get_or_create(
        contrato=n_mppe,
        defaults={
            'supplier_contract':      supplier,
            'numero_contrato_oficial': (header.get('numero_contrato') or '')[:30],
            'forma':                   resumo_data.get('forma', ''),
            'numero_processo':         (resumo_data.get('numero_processo') or '')[:60],
            'numero_arp':              (resumo_data.get('numero_arp') or '')[:60],
            'categoria':               (resumo_data.get('categoria') or '')[:60],
            'aditivo_prazo':           (resumo_data.get('aditivo_prazo') or '')[:100],
            'aditivo_preco':           (resumo_data.get('aditivo_preco') or '')[:100],
            'valor_contratado':        resumo_data.get('valor_contratado'),
            'previsao_anual':          resumo_data.get('previsao_anual'),
            'inicio_vigencia':         header.get('inicio_vigencia'),
            'final_vigencia':          header.get('final_vigencia') or (
                                          resumo_data['vencimento'].date()
                                          if resumo_data.get('vencimento') else None
                                       ),
            'homologacao':             header.get('homologacao'),
            'numero_sc':               (header.get('numero_sc') or '')[:30],
            'cod_liquidacao':          header.get('cod_liquidacao'),
        }
    )

    if not created and sobrescrever:
        update_fields = []
        if supplier:
            contrato.supplier_contract = supplier
            update_fields.append('supplier_contract')
        for field, value in [
            ('valor_contratado',  resumo_data.get('valor_contratado')),
            ('previsao_anual',    resumo_data.get('previsao_anual')),
            ('categoria',         (resumo_data.get('categoria') or '')[:60]),
            ('forma',             resumo_data.get('forma', '')),
            ('numero_processo',   (resumo_data.get('numero_processo') or '')[:60]),
            ('aditivo_prazo',     (resumo_data.get('aditivo_prazo') or '')[:100]),
            ('aditivo_preco',     (resumo_data.get('aditivo_preco') or '')[:100]),
            ('inicio_vigencia',   header.get('inicio_vigencia')),
            ('final_vigencia',    header.get('final_vigencia')),
            ('homologacao',       header.get('homologacao')),
        ]:
            if value is not None:
                setattr(contrato, field, value)
                update_fields.append(field)
        if update_fields:
            contrato.save(update_fields=update_fields)

    return contrato, created


def _upsert_saldo_ativo(contrato, item, sobrescrever):
    """Cria ou atualiza SaldoAtivo. Retorna (obj, criado, motivo_skip)."""
    efisco_str = item['efisco']
    try:
        bens_consumo = BensConsumo.objects.get(efisco=efisco_str)
    except BensConsumo.DoesNotExist:
        return None, False, f'E-Fisco {efisco_str} não encontrado no cadastro'

    qtd = item['quantidade']
    if qtd is None or qtd <= 0:
        return None, False, f'E-Fisco {efisco_str}: quantidade inválida ({qtd})'

    saldo, created = SaldoAtivo.objects.get_or_create(
        contrato_saldo=contrato,
        efisco=bens_consumo,
        defaults={
            'marca':               item['marca'][:60] if item['marca'] else '',
            'descricao_manual':    item['descricao'][:60] if item['descricao'] else '',
            'quantidade_contrato': qtd,
            'numero_item':         item['numero_item'],
            'preco_unitario':      item['preco_unit'],
            # saldo_disponivel é preenchido automaticamente pelo save()
        }
    )

    if not created and sobrescrever:
        saldo.quantidade_contrato = qtd
        saldo.numero_item         = item['numero_item']
        saldo.preco_unitario      = item['preco_unit']
        if item['marca']:
            saldo.marca = item['marca'][:60]
        if item['descricao']:
            saldo.descricao_manual = item['descricao'][:60]
        # Atualiza saldo disponível — clampeado em 0 (overexecution vira 0)
        if item.get('qtd_saldo') is not None:
            saldo.saldo_disponivel = max(0, int(item['qtd_saldo']))
        saldo.save()

    return saldo, created, None


# ─── Command principal ──────────────────────────────────────────────────────

class Command(BaseCommand):
    help = 'Importa saldo ativo de planilha XLSX da DIMMS'

    def add_arguments(self, parser):
        parser.add_argument(
            'xlsx',
            nargs='?',
            default='docs/Saldo ativo de materiais 2026.xlsx',
            help='Caminho para o arquivo XLSX (padrão: docs/Saldo ativo de materiais 2026.xlsx)',
        )
        parser.add_argument(
            '--sobrescrever',
            action='store_true',
            help='Atualiza registros existentes (fornecedor, contrato, itens)',
        )

    def handle(self, *args, **options):
        xlsx_path  = options['xlsx']
        sobrescrever = options['sobrescrever']

        self.stdout.write(f'Abrindo {xlsx_path}...')
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'Arquivo não encontrado: {xlsx_path}')

        resumo_map = _build_resumo_map(wb)
        self.stdout.write(f'RESUMO: {len(resumo_map)} contratos mapeados.')

        # Abas a ignorar (não têm itens no formato padrão)
        SKIP_SHEETS = {
            'RESUMO', 'EMPENHOS 2025', 'EMPENHOS 2026', 'LISTAGEM',
            'NEs Anulação', 'NEs Anula??o', 'resumo pedido',
            'Resumo Papel Sulfite A4', 'Detalhamento de consumo papel A',
            'Página51', 'P?gina51',
        }

        totais = {'sheets': 0, 'contratos_criados': 0, 'contratos_atualizados': 0,
                  'itens_criados': 0, 'itens_atualizados': 0, 'itens_ignorados': 0}

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws   = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            items = _parse_items(rows)
            if not items:
                continue   # aba sem dados de item no formato esperado

            totais['sheets'] += 1
            header = _parse_header(rows)

            # Normaliza o nome da aba para buscar no mapa do RESUMO
            sheet_norm  = _normalize(sheet_name)
            resumo_data = resumo_map.get(sheet_norm, {})

            # N° MPPE: usa o nome da aba (é o identificador interno)
            n_mppe = sheet_name.strip()

            fornecedor_nome = (
                header.get('fornecedor') or resumo_data.get('fornecedor') or ''
            )

            with transaction.atomic():
                supplier = _get_or_create_supplier(header, fornecedor_nome, sobrescrever)
                contrato, c_created = _get_or_create_contrato(
                    n_mppe, header, resumo_data, supplier, sobrescrever
                )

            if c_created:
                totais['contratos_criados'] += 1
                self.stdout.write(f'  [CONTRATO +] {n_mppe}')
            else:
                totais['contratos_atualizados'] += 1

            for item in items:
                with transaction.atomic():
                    obj, i_created, motivo = _upsert_saldo_ativo(contrato, item, sobrescrever)

                if motivo:
                    self.stderr.write(f'    [SKIP] {motivo}')
                    totais['itens_ignorados'] += 1
                elif i_created:
                    totais['itens_criados'] += 1
                else:
                    totais['itens_atualizados'] += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nConcluido!\n'
            f'  Abas processadas : {totais["sheets"]}\n'
            f'  Contratos criados: {totais["contratos_criados"]}\n'
            f'  Contratos atualizados: {totais["contratos_atualizados"]}\n'
            f'  Itens criados    : {totais["itens_criados"]}\n'
            f'  Itens atualizados: {totais["itens_atualizados"]}\n'
            f'  Itens ignorados  : {totais["itens_ignorados"]}'
        ))
