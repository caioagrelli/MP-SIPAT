"""
Atualiza o estoque de bens de consumo a partir de uma aba de planilha .xlsx.

Suporta dois formatos de coluna detectados automaticamente pelo cabeçalho:

Formato A — Planilha1 (importação original, 15 colunas):
  0  ESSENCIAL  1  Índice  2  Almoxarifado  3  Grupo  4  E-fisco
  5  Descrição  6  Unidade  8  Marca  10  Consumo mensal
  13  Validade  14  Qtd. Estoque

Formato B — Planilha2 (inventário atualizado, 12 colunas):
  0  Almoxarifado  1  Grupo  2  E-fisco  3  Descrição  4  Unidade
  9  Marca nova    10  Validade nova    11  Qtd. Inventariada

Comportamento:
  - Item já existe  → atualiza amount_shock, validity (Estoque)
                      e almoxarifado (BensConsumo).
  - Item novo       → cria BensConsumo + Estoque.

Uso:
    python manage.py atualizar_estoque_xlsx <arquivo.xlsx>
    python manage.py atualizar_estoque_xlsx <arquivo.xlsx> --aba "Planilha2"
"""
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from dimms.models import BensConsumo, Estoque
from dimms.utils import GrupoConsumo, TipoAlmoxarifado, UnidadesMedida


# ---------------------------------------------------------------------------
# Normalização / mapeamento
# ---------------------------------------------------------------------------

def _norm(texto: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFD', str(texto))
        if unicodedata.category(c) != 'Mn'
    ).lower().strip()


_GRUPO_MAP = {
    'alimentos':                     GrupoConsumo.alimento,
    'alimento':                      GrupoConsumo.alimento,
    'confeccao':                     GrupoConsumo.confeccao,
    'copa-cozinha':                  GrupoConsumo.copa_cozinha,
    'copa cozinha':                  GrupoConsumo.copa_cozinha,
    'domissanitarios':               GrupoConsumo.domissanitario,
    'domissanitario':                GrupoConsumo.domissanitario,
    'eletrica':                      GrupoConsumo.eletrica,
    'epi':                           GrupoConsumo.epi,
    'hidrossanitario':               GrupoConsumo.hidrosanitario,
    'hidrosanitario':                GrupoConsumo.hidrosanitario,
    'informatica':                   GrupoConsumo.informatica,
    'limpeza':                       GrupoConsumo.limpeza,
    'manutencao comum':              GrupoConsumo.manutencao,
    'manutencao':                    GrupoConsumo.manutencao,
    'marcenaria':                    GrupoConsumo.marcenaria,
    'pintura':                       GrupoConsumo.pintura,
    'papeis para expediente':        GrupoConsumo.papeis_expediente,
    'papeis de expediente':          GrupoConsumo.papeis_expediente,
    'expediente':                    GrupoConsumo.papeis_expediente,
    'papeis para limpeza':           GrupoConsumo.papeis_limpeza,
    'papeis de limpeza':             GrupoConsumo.papeis_limpeza,
    'refrigeracao':                  GrupoConsumo.refrigeracao,
    'toners':                        GrupoConsumo.toner,
    'toner':                         GrupoConsumo.toner,
    'construcao civil':              GrupoConsumo.manutencao,
    'bens permanentes transferidos': GrupoConsumo.manutencao,
    'estocagem':                     GrupoConsumo.manutencao,
    'telecom':                       GrupoConsumo.informatica,
}


def _mapear_grupo(raw) -> str:
    if not raw:
        return GrupoConsumo.manutencao
    return _GRUPO_MAP.get(_norm(str(raw)), GrupoConsumo.manutencao)


def _mapear_unidade(raw) -> str:
    if not raw:
        return UnidadesMedida.unidade
    s = _norm(str(raw))
    if s.startswith('quilograma') or s in ('kg',) or s.startswith('sc'):
        return UnidadesMedida.quilograma
    if s.startswith('litro') or s.startswith('lt') or s.startswith('gl') or s == 'bombona' or s.startswith('fs'):
        return UnidadesMedida.litro
    if s.startswith('pacote') or s in ('pct', 'pct.', 'pc'):
        return UnidadesMedida.pacote
    if s.startswith('caixa') or s in ('cx',) or s.startswith('crt'):
        return UnidadesMedida.caixa
    if s in ('m', 'm2', 'm²', 'rolo') or s.startswith('rl') or s.startswith('vara') or 'metro' in s:
        return UnidadesMedida.metro
    return UnidadesMedida.unidade


def _mapear_almoxarifado(raw) -> str:
    if not raw:
        return TipoAlmoxarifado.geral
    s = _norm(str(raw))
    if 'reserv' in s:
        return TipoAlmoxarifado.reservado
    return TipoAlmoxarifado.geral


# ---------------------------------------------------------------------------
# Helpers de célula
# ---------------------------------------------------------------------------

def _cel(row, idx: int) -> str:
    try:
        val = row[idx].value
    except IndexError:
        return ''
    return '' if val is None else str(val).strip()


def _inteiro(val) -> int:
    if val is None or val == '':
        return 0
    try:
        return max(0, int(float(str(val))))
    except (ValueError, TypeError):
        return 0


def _date(val):
    if val is None or val == '':
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Detectar formato da planilha pelo cabeçalho
# ---------------------------------------------------------------------------

def _detectar_formato(header_row) -> str:
    """Retorna 'A' (formato original) ou 'B' (inventário atualizado)."""
    valores = [str(c.value or '').strip().upper() for c in header_row]
    # Formato B tem 'E-FISCO' na posição 2 e 'QTD. INVENTARIADA' na 11
    if len(valores) > 2 and 'FISCO' in valores[2]:
        return 'B'
    return 'A'


# ---------------------------------------------------------------------------
# Processar linha
# ---------------------------------------------------------------------------

def _process_row(row, formato: str) -> dict:
    if formato == 'B':
        # Formato B: 0=Almox, 1=Grupo, 2=Efisco, 3=Desc, 4=Unidade,
        #            9=Marca nova, 10=Validade nova, 11=Qtd inventariada
        efisco_raw   = _cel(row, 2)
        efisco       = re.sub(r'\.0$', '', efisco_raw)[:20]
        descricao    = _cel(row, 3) or efisco
        grupo        = _mapear_grupo(_cel(row, 1))
        unidade      = _mapear_unidade(_cel(row, 4))
        almoxarifado = _mapear_almoxarifado(_cel(row, 0))
        marca        = (_cel(row, 9) or _cel(row, 5) or 'Não informada')[:40]
        validade     = _date(row[10].value)
        qtd          = _inteiro(row[11].value)
        consumo      = None
        essencial    = False
    else:
        # Formato A: estrutura original
        efisco_raw   = _cel(row, 4)
        efisco       = re.sub(r'\.0$', '', efisco_raw)[:20]
        descricao    = _cel(row, 5) or efisco
        grupo        = _mapear_grupo(_cel(row, 3))
        unidade      = _mapear_unidade(_cel(row, 6))
        almoxarifado = _mapear_almoxarifado(_cel(row, 2))
        marca        = (_cel(row, 8) or 'Não informada')[:40]
        consumo      = _inteiro(row[10].value)
        validade     = _date(row[13].value)
        qtd          = _inteiro(row[14].value)
        essencial    = _cel(row, 0).lower() in ('sim', 's', '1', 'true', 'x')

    if not efisco or efisco.lower() in ('none', ''):
        return {'status': 'ignorado', 'efisco': '—', 'msg': 'Sem código E-fisco'}

    try:
        with transaction.atomic():
            bem = BensConsumo.objects.filter(efisco=efisco).first()

            if bem:
                # --- Atualiza BensConsumo ---
                bem.almoxarifado = almoxarifado
                bem.save(update_fields=['almoxarifado'])

                # --- Atualiza Estoque mais recente ---
                estoque = bem.bem_estoque.order_by('-created_at').first()
                if estoque:
                    estoque.amount_shock = qtd
                    estoque.validity     = validade
                    estoque.essential    = essencial
                    fields = ['amount_shock', 'validity', 'essential']
                    # Só atualiza consumo mensal se a planilha trouxer o valor
                    if consumo is not None:
                        estoque.monthly_consumption = consumo if consumo else None
                        fields.append('monthly_consumption')
                    estoque.save(update_fields=fields)
                else:
                    # Bem sem estoque associado — cria
                    Estoque.objects.create(
                        item_shock=bem,
                        description_manual=descricao[:90],
                        mark=marca,
                        amount_shock=qtd,
                        monthly_consumption=consumo if consumo else None,
                        essential=essencial,
                        validity=validade,
                        form_input='Importação',
                    )

                return {'status': 'atualizado', 'efisco': efisco, 'msg': 'Atualizado'}

            else:
                # --- Cria novo ---
                bem = BensConsumo.objects.create(
                    efisco=efisco,
                    descricao_efisco=descricao,
                    medida=unidade,
                    grupo_consumo=grupo,
                    almoxarifado=almoxarifado,
                )
                Estoque.objects.create(
                    item_shock=bem,
                    description_manual=descricao[:90],
                    mark=marca,
                    amount_shock=qtd,
                    monthly_consumption=consumo if consumo else None,
                    essential=essencial,
                    validity=validade,
                    form_input='Importação',
                )
                return {'status': 'criado', 'efisco': efisco, 'msg': 'Criado'}

    except Exception as exc:
        return {'status': 'erro', 'efisco': efisco, 'msg': str(exc)}


# ---------------------------------------------------------------------------
# Management command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = 'Atualiza estoque de bens de consumo a partir de aba de planilha .xlsx'

    def add_arguments(self, parser):
        parser.add_argument('arquivo', type=str, help='Caminho para o arquivo .xlsx')
        parser.add_argument(
            '--aba',
            type=str,
            default=None,
            help='Nome da aba a ler (padrão: última aba do arquivo)',
        )

    def handle(self, *args, **options):
        caminho = Path(options['arquivo'])
        if not caminho.exists():
            raise CommandError(f'Arquivo não encontrado: {caminho}')

        self.stdout.write(f'Abrindo {caminho.name}...')
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

        nome_aba = options['aba']
        if nome_aba:
            if nome_aba not in wb.sheetnames:
                raise CommandError(
                    f'Aba "{nome_aba}" não encontrada. Abas disponíveis: {wb.sheetnames}'
                )
            ws = wb[nome_aba]
        else:
            # Última aba
            ws = wb[wb.sheetnames[-1]]
            nome_aba = wb.sheetnames[-1]

        self.stdout.write(f'Lendo aba: "{nome_aba}"')

        rows = list(ws.iter_rows())
        if not rows:
            raise CommandError('Aba vazia.')

        formato = _detectar_formato(rows[0])
        self.stdout.write(f'Formato detectado: {"B (inventário)" if formato == "B" else "A (original)"}')

        # Formato A: cabeçalho linha 0, vazia linha 1, dados a partir de 2
        # Formato B: cabeçalho linha 0, dados a partir de 1
        inicio = 2 if formato == 'A' else 1
        dados = [r for r in rows[inicio:] if any(c.value is not None for c in r)]
        total = len(dados)

        if total == 0:
            raise CommandError('Aba sem dados.')

        self.stdout.write(f'Linhas com dados: {total}')

        atualizados = 0
        criados     = 0
        ignorados   = 0
        erros       = []

        for i, row in enumerate(dados, 1):
            res = _process_row(row, formato)

            if res['status'] == 'atualizado':
                atualizados += 1
            elif res['status'] == 'criado':
                criados += 1
            elif res['status'] == 'ignorado':
                ignorados += 1
            else:
                erros.append(res)

            if i % 100 == 0:
                pct = i * 100 // total
                self.stdout.write(
                    f'  [{pct:3d}%] {i}/{total} — '
                    f'atualizados: {atualizados}  criados: {criados}  erros: {len(erros)}',
                    ending='\r',
                )
                self.stdout.flush()

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'Atualizados: {atualizados}'))
        self.stdout.write(self.style.SUCCESS(f'Criados:     {criados}'))
        if ignorados:
            self.stdout.write(f'Ignorados:   {ignorados}')
        if erros:
            self.stdout.write(self.style.ERROR(f'Erros:       {len(erros)}'))
            for e in erros:
                self.stdout.write(f'  E-fisco {e["efisco"]}: {e["msg"]}')
        else:
            self.stdout.write(self.style.SUCCESS('Sem erros!'))
