# bibliotecas do django
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# importações do código
from ..forms import InfoUAForm
from ..models import InfoUA, CircunscricaoPredio
from dimms.models import Solicitacao

# =========================================
# VIEWS DAS UAS (UNIDADES ADMINISTRATIVAS)
# =========================================



''' Unidades Administrativas '''
# Página inicial das UAs
@login_required
def ua_homepage(request):
    query = request.GET.get("q", "").strip()
    predio_filter = request.GET.get("predio", "").strip()
    sede_filter = request.GET.get("sede", "").strip()

    total_cadastradas = InfoUA.objects.count()
    total_sedes = InfoUA.objects.filter(sede=True).count()
    total_predios = CircunscricaoPredio.objects.count()

    uas = InfoUA.objects.select_related("circunscricao_predio").all().order_by("ua")

    if query:
        uas = uas.filter(
            Q(ua__icontains=query) |
            Q(responsavel_ua__icontains=query) |
            Q(email_ua__icontains=query) |
            Q(circunscricao_predio__local__icontains=query)
        )

    if predio_filter:
        uas = uas.filter(circunscricao_predio__pk=predio_filter)

    if sede_filter == "1":
        uas = uas.filter(sede=True)
    elif sede_filter == "0":
        uas = uas.filter(sede=False)

    predios = CircunscricaoPredio.objects.order_by("local")

    context = {
        "uas": uas,
        "query": query,
        "predio_filter": predio_filter,
        "sede_filter": sede_filter,
        "predios": predios,
        "total_cadastradas": total_cadastradas,
        "total_sedes": total_sedes,
        "total_predios": total_predios,
    }
    return render(request, "dempam/uas/ua_homepage.html", context)


# Planilha (.xlsx) com todas as UAs cadastradas, pra preencher/conferir informações
@login_required
def ua_export_xlsx(request):
    uas = (
        InfoUA.objects
        .select_related("circunscricao_predio", "gestor")
        .order_by("ua")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "UAs"

    colunas = [
        ("Código", 14),
        ("UA", 46),
        ("Sigla", 14),
        ("Circunscrição/Prédio", 32),
        ("Nível", 8),
        ("Sede", 8),
        ("Responsável da UA", 30),
        ("Matrícula do Responsável", 20),
        ("Contato da UA", 16),
        ("Email da UA", 32),
        ("Gestor (usuário)", 22),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E2D42", end_color="1E2D42", fill_type="solid")

    for col_idx, (titulo, largura) in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=col_idx, value=titulo)
        cel.font = header_font
        cel.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = "A2"

    for row_idx, ua in enumerate(uas, start=2):
        ws.cell(row=row_idx, column=1, value=ua.codigo or "")
        ws.cell(row=row_idx, column=2, value=ua.ua)
        ws.cell(row=row_idx, column=3, value=ua.sigla or "")
        ws.cell(row=row_idx, column=4, value=ua.circunscricao_predio.local if ua.circunscricao_predio else "")
        ws.cell(row=row_idx, column=5, value=ua.nivel)
        ws.cell(row=row_idx, column=6, value="Sim" if ua.sede else "Não")
        ws.cell(row=row_idx, column=7, value=ua.responsavel_ua or "")
        ws.cell(row=row_idx, column=8, value=ua.mat_resp_ua)
        ws.cell(row=row_idx, column=9, value=ua.contato_ua or "")
        ws.cell(row=row_idx, column=10, value=ua.email_ua or "")
        ws.cell(row=row_idx, column=11, value=ua.gestor.get_full_name() or ua.gestor.username if ua.gestor else "")

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="uas_sipat.xlsx"'
    wb.save(resp)
    return resp


@login_required
def ua_detail(request, pk):
    ua = get_object_or_404(InfoUA.objects.select_related('circunscricao_predio'), pk=pk)
    solicitacoes = ua.solicitantesconsumo.select_related('user_responsible').order_by('-data_order')

    total = solicitacoes.count()
    em_atendimento = solicitacoes.filter(situation='ATENDIMENTO').count()
    recebidas = solicitacoes.filter(situation='RECEBIDA').count()
    canceladas = solicitacoes.filter(situation='CANCELADA').count()

    return render(request, 'dempam/uas/ua_detail.html', {
        'ua': ua,
        'solicitacoes': solicitacoes,
        'total': total,
        'em_atendimento': em_atendimento,
        'recebidas': recebidas,
        'canceladas': canceladas,
    })


# Cadastrar nova UA
@login_required
def ua_add(request):
    if request.method == "POST":
        form = InfoUAForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "UA cadastrada com sucesso.")
            return redirect("dempam:ua_homepage")
    else:
        form = InfoUAForm()

    context = {
        "form": form,
        "page_title": "Cadastrar UA",
        "button_label": "Salvar UA",
    }
    return render(request, "dempam/uas/ua_add.html", context)

@login_required
def ua_update(request, pk):
    ua = get_object_or_404(InfoUA, pk=pk)

    if request.method == "POST":
        form = InfoUAForm(request.POST, instance=ua)
        if form.is_valid():
            form.save()
            messages.success(request, "UA atualizada com sucesso.")
            return redirect("dempam:ua_detail", pk=pk)
    else:
        form = InfoUAForm(instance=ua)

    return render(request, "dempam/uas/ua_update.html", {
        "form": form,
        "ua_obj": ua,
        "page_title": "Editar UA",
        "button_label": "Atualizar UA",
    })

# Busca de UA por nome (autocomplete)
@login_required
def ua_search(request):
    from django.http import JsonResponse
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse([], safe=False)
    uas = (
        InfoUA.objects
        .filter(Q(ua__icontains=q) | Q(responsavel_ua__icontains=q))
        .values('id', 'ua', 'responsavel_ua')
        [:20]
    )
    return JsonResponse(list(uas), safe=False)
