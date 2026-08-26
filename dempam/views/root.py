# bibliotecas do django
import re
from functools import wraps
from decimal import Decimal
from collections import defaultdict
from types import SimpleNamespace

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from django.conf import settings
from django.utils import timezone
from django.db.models import Max, Count, Sum, F, Q
from django.http import JsonResponse, HttpResponse, QueryDict
from django.core.paginator import Paginator

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# importações do código
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST

from ..models import InfoUA, Aviso, Municipio
from ..forms import AvisoForm
from ..services import montar_dashboard_usuario
from dimms.models import Estoque, Solicitacao, SolicitacaoItens
from dimrcbp.models import BensPermanentes
from dimms.views.relatorios import _montar_gastos_por_item, MESES_PT
from dimms.utils import GrupoConsumo

# ====================================================
# VIEWS CENTRAIS DO SIPAT (DIRECIONAMENTO DE PÁGINAS)
# ====================================================


def gerencia_dempam_required(view_func):
    """Permite superusuários, staff, ou membros da Gerência DEMPAM — mais restrito
    que management_required (que também libera gestão de usuários/grupos)."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        if (
            user.is_superuser
            or user.is_staff
            or user.groups.filter(name='Gerência DEMPAM').exists()
        ):
            return view_func(request, *args, **kwargs)
        raise PermissionDenied
    return wrapper



''' FUNÇÕES PRINCIPAIS DO SIPAT'''
# destinar paginas
def root(request):
    if request.user.is_authenticated:
        return redirect('home')  # Página quando o usuário estiver logado
    next_url = request.GET.get('next', '')
    login_url = f"{settings.LOGIN_URL}?next={next_url}" if next_url else settings.LOGIN_URL
    return redirect(login_url)

#homepage central (futuramente vai ser um app a parte)
@login_required
def home(request):
    context = montar_dashboard_usuario(request.user)
    return render(request, 'global/home.html', context)
  
  
''' Homepage'''
# pagina principal do dempam
@login_required
def homepage(request):
    municipios_com_circunscricao = Municipio.objects.exclude(circunscricao='').order_by('nome')
    circunscricao_por_municipio = {
        m.codigo_ibge: m.circunscricao for m in municipios_com_circunscricao
    }

    def _ordem_circunscricao(label):
        m = re.match(r'^(\d+)ª', label)
        return (0, int(m.group(1))) if m else (1, label)

    circunscricoes = sorted(set(circunscricao_por_municipio.values()), key=_ordem_circunscricao)

    return render(request, 'dempam/homepage.html', {
        'total_setores': InfoUA.objects.count(),
        'total_bens_permanentes': BensPermanentes.objects.count(),
        'total_consumo': Estoque.objects.count(),
        'total_pendencias': Solicitacao.objects.filter(situation='ATENDIMENTO').count(),
        'circunscricao_por_municipio': circunscricao_por_municipio,
        'circunscricoes': circunscricoes,
    })


''' Mapa do Painel Gerencial '''
# Monta o resumo (UAs + bens + pedidos + gasto) de um mês específico pra um conjunto
# de UAs — usado tanto pro clique num município quanto pra seleção de uma
# circunscrição inteira. Sem parâmetro de mês, usa o mês mais recente com solicitação.
def _montar_resumo_uas(nome, uas, mes_param=None):
    total_bens = BensPermanentes.objects.filter(history_tombo__current_ua__in=uas).count()

    solicitacoes = Solicitacao.objects.filter(ua_order__in=uas)

    meses_disponiveis = [
        {'ano': m['data_order__year'], 'mes': m['data_order__month']}
        for m in (
            solicitacoes
            .exclude(data_order__isnull=True)
            .values('data_order__year', 'data_order__month')
            .distinct()
            .order_by('-data_order__year', '-data_order__month')
        )
    ]

    ano = mes = None
    if mes_param:
        try:
            ano_str, mes_str = mes_param.split('-')
            candidato = (int(ano_str), int(mes_str))
        except (ValueError, AttributeError):
            candidato = None
        if candidato and any((m['ano'], m['mes']) == candidato for m in meses_disponiveis):
            ano, mes = candidato
    if ano is None and meses_disponiveis:
        ano, mes = meses_disponiveis[0]['ano'], meses_disponiveis[0]['mes']

    total_pedidos_mes = 0
    total_gasto_mes = Decimal('0')
    if ano and mes:
        solicitacoes_mes = solicitacoes.filter(data_order__year=ano, data_order__month=mes)
        total_pedidos_mes = solicitacoes_mes.count()
        itens = (
            SolicitacaoItens.objects
            .filter(
                request_defendant__in=solicitacoes_mes,
                item_order__isnull=False,
                amount_order__isnull=False,
            )
            .select_related('item_order__item_shock')
        )
        for item in itens:
            preco = item.item_order.item_shock.preco_medio
            if preco:
                total_gasto_mes += item.amount_order * preco

    return {
        'nome': nome,
        'uas': list(uas.order_by('ua').values_list('ua', 'sigla')),
        'total_bens': total_bens,
        'total_pedidos': total_pedidos_mes,
        'gasto_mensal': float(total_gasto_mes),
        'mes_label': f'{MESES_PT[mes]}/{ano}' if mes else None,
        'mes_selecionado': f'{ano}-{mes:02d}' if mes else None,
        'meses_disponiveis': [
            {
                'valor': f"{m['ano']}-{m['mes']:02d}",
                'label': f"{MESES_PT[m['mes']][:3]}/{str(m['ano'])[2:]}",
            }
            for m in meses_disponiveis
        ],
    }


# Resumo das UAs de um único município, pro clique no mapa
@login_required
def municipio_resumo(request, codigo_ibge):
    municipio = get_object_or_404(Municipio, codigo_ibge=codigo_ibge)
    uas = InfoUA.objects.filter(municipio=municipio)
    return JsonResponse(_montar_resumo_uas(municipio.nome, uas, request.GET.get('mes')))


# Resumo agregado de todas as UAs dos municípios de uma circunscrição, pro filtro do mapa
@login_required
def circunscricao_resumo(request, circunscricao):
    uas = InfoUA.objects.filter(municipio__circunscricao=circunscricao)
    return JsonResponse(_montar_resumo_uas(circunscricao, uas, request.GET.get('mes')))


# Ranking de gasto (consumo) por município e por circunscrição — "gasto mensal"
# é o gasto real do mês da última solicitação de cada chave (não uma média
# histórica), pra não diluir picos/quedas recentes com meses antigos.
def _linha_ranking_vazia():
    return {'total_bens': 0, 'total_pedidos': 0, 'gasto_ultimo_mes': Decimal('0'), 'ultima': None}


def _finalizar_ranking(dados_por_chave, nomes):
    linhas = [
        {
            'nome': nomes[chave],
            'total_bens': dados['total_bens'],
            'total_pedidos': dados['total_pedidos'],
            'gasto_mensal': dados['gasto_ultimo_mes'],
        }
        for chave, dados in dados_por_chave.items()
    ]
    linhas.sort(key=lambda l: l['gasto_mensal'], reverse=True)
    return linhas


# Monta o ranking (por município ou por circunscrição) agregando direto no banco
# (GROUP BY), em vez de iterar linha a linha em Python — necessário porque
# BensPermanentes sozinho já passa de 40 mil registros.
def _montar_ranking(campo_agrupamento, filtro_extra=None):
    dados = defaultdict(_linha_ranking_vazia)
    nomes = {}

    solicitacoes = (
        Solicitacao.objects
        .filter(
            **{f'ua_order__{campo_agrupamento}__isnull': False},
            **{f'ua_order__{k}': v for k, v in (filtro_extra or {}).items()},
        )
        .values(f'ua_order__{campo_agrupamento}')
        .annotate(total=Count('id'), ultima=Max('data_order'))
    )
    for linha in solicitacoes:
        chave = linha[f'ua_order__{campo_agrupamento}']
        dados[chave]['total_pedidos'] = linha['total']
        dados[chave]['ultima'] = linha['ultima']

    # gasto agrupado por (chave, ano, mês) — depois ficamos só com o mês da
    # última solicitação de cada chave, em vez de somar/dividir tudo
    itens_por_mes = (
        SolicitacaoItens.objects
        .filter(
            item_order__isnull=False,
            amount_order__isnull=False,
            item_order__item_shock__preco_medio__isnull=False,
            **{f'request_defendant__ua_order__{campo_agrupamento}__isnull': False},
            **{f'request_defendant__ua_order__{k}': v for k, v in (filtro_extra or {}).items()},
        )
        .annotate(custo=F('amount_order') * F('item_order__item_shock__preco_medio'))
        .values(
            f'request_defendant__ua_order__{campo_agrupamento}',
            'request_defendant__data_order__year',
            'request_defendant__data_order__month',
        )
        .annotate(total_gasto=Sum('custo'))
    )
    gasto_por_chave_mes = {}
    for linha in itens_por_mes:
        chave = linha[f'request_defendant__ua_order__{campo_agrupamento}']
        ano_mes = (linha['request_defendant__data_order__year'], linha['request_defendant__data_order__month'])
        gasto_por_chave_mes[(chave, ano_mes)] = linha['total_gasto'] or Decimal('0')

    for chave, d in dados.items():
        if d['ultima']:
            ano_mes = (d['ultima'].year, d['ultima'].month)
            d['gasto_ultimo_mes'] = gasto_por_chave_mes.get((chave, ano_mes), Decimal('0'))

    bens = (
        BensPermanentes.objects
        .filter(
            **{f'history_tombo__current_ua__{campo_agrupamento}__isnull': False},
            **{f'history_tombo__current_ua__{k}': v for k, v in (filtro_extra or {}).items()},
        )
        .values(f'history_tombo__current_ua__{campo_agrupamento}')
        .annotate(total=Count('id'))
    )
    for linha in bens:
        chave = linha[f'history_tombo__current_ua__{campo_agrupamento}']
        dados[chave]['total_bens'] = linha['total']

    if campo_agrupamento == 'municipio':
        for m in Municipio.objects.filter(pk__in=dados.keys()):
            nomes[m.pk] = m.nome
    else:
        nomes = {chave: chave for chave in dados.keys()}

    return _finalizar_ranking(dados, nomes)


@login_required
def ranking_gastos(request):
    ranking_municipios = _montar_ranking('municipio')
    ranking_circunscricoes = _montar_ranking('municipio__circunscricao', filtro_extra={'municipio__circunscricao__gt': ''})

    paginator = Paginator(ranking_municipios, 15)
    pagina_municipios = paginator.get_page(request.GET.get('pagina'))

    return render(request, 'dempam/ranking_gastos.html', {
        'pagina_municipios': pagina_municipios,
        'ranking_circunscricoes': ranking_circunscricoes,
    })


GASTOS_UA_POR_PAGINA = 30


# Agrega SolicitacaoItens por UA (gasto = quantidade × preço médio) — usada tanto
# pela tela do relatório (paginada) quanto pela exportação XLSX (sem paginar)
def _montar_gastos_por_ua(request):
    data_de = request.GET.get('data_de', '').strip()
    data_ate = request.GET.get('data_ate', '').strip()
    busca = request.GET.get('busca', '').strip()

    solicitacoes = Solicitacao.objects.filter(ua_order__isnull=False)
    itens = SolicitacaoItens.objects.filter(
        item_order__isnull=False,
        amount_order__isnull=False,
        request_defendant__ua_order__isnull=False,
    )
    if data_de:
        solicitacoes = solicitacoes.filter(data_order__date__gte=data_de)
        itens = itens.filter(request_defendant__data_order__date__gte=data_de)
    if data_ate:
        solicitacoes = solicitacoes.filter(data_order__date__lte=data_ate)
        itens = itens.filter(request_defendant__data_order__date__lte=data_ate)

    dados = defaultdict(lambda: {'total_pedidos': 0, 'total_qtd': Decimal('0'), 'total_gasto': Decimal('0')})

    for linha in solicitacoes.values('ua_order').annotate(total=Count('id')):
        dados[linha['ua_order']]['total_pedidos'] = linha['total']

    for linha in (
        itens
        .annotate(custo=F('amount_order') * F('item_order__item_shock__preco_medio'))
        .values('request_defendant__ua_order')
        .annotate(total_gasto=Sum('custo'), total_qtd=Sum('amount_order'))
    ):
        chave = linha['request_defendant__ua_order']
        dados[chave]['total_gasto'] = linha['total_gasto'] or Decimal('0')
        dados[chave]['total_qtd'] = linha['total_qtd'] or Decimal('0')

    qtd_itens_sem_preco = (
        itens.filter(item_order__item_shock__preco_medio__isnull=True)
        .values('item_order__item_shock').distinct().count()
    )

    uas = InfoUA.objects.filter(pk__in=dados.keys()).select_related('municipio')
    if busca:
        uas = uas.filter(Q(ua__icontains=busca) | Q(sigla__icontains=busca))

    linhas = [
        {
            'ua': ua,
            'total_pedidos': dados[ua.pk]['total_pedidos'],
            'total_qtd': dados[ua.pk]['total_qtd'],
            'total_gasto': dados[ua.pk]['total_gasto'],
        }
        for ua in uas
    ]
    linhas.sort(key=lambda l: l['total_gasto'], reverse=True)

    total_geral = sum((l['total_gasto'] for l in linhas), Decimal('0'))
    total_qtd_geral = sum((l['total_qtd'] for l in linhas), Decimal('0'))

    return {
        'linhas': linhas,
        'data_de': data_de,
        'data_ate': data_ate,
        'busca': busca,
        'total_geral': total_geral,
        'total_qtd_geral': total_qtd_geral,
        'qtd_itens_sem_preco': qtd_itens_sem_preco,
    }


''' Gastos por UA — quantidade saída × preço médio, agrupado por UA (com município/IBGE pra cruzar com o mapa) '''
@login_required
def gastos_por_ua(request):
    dados = _montar_gastos_por_ua(request)
    pagina = Paginator(dados['linhas'], GASTOS_UA_POR_PAGINA).get_page(request.GET.get('pagina'))
    return render(request, 'dempam/gastos_por_ua.html', {**dados, 'pagina': pagina})


''' Gastos por UA — exportação XLSX (linha completa, respeitando os filtros da tela) para uso externo (ex.: Power BI) '''
@login_required
def gastos_por_ua_export_xlsx(request):
    dados = _montar_gastos_por_ua(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gastos por UA'

    colunas = [
        ('UA', 46),
        ('Sigla', 14),
        ('Município', 28),
        ('Código IBGE', 14),
        ('Circunscrição', 26),
        ('Solicitações', 14),
        ('Quantidade', 14),
        ('Gasto', 14),
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E2D42', end_color='1E2D42', fill_type='solid')

    for col_idx, (titulo, largura) in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=col_idx, value=titulo)
        cel.font = header_font
        cel.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = 'A2'

    for row_idx, linha in enumerate(dados['linhas'], start=2):
        ua = linha['ua']
        ws.cell(row=row_idx, column=1, value=ua.ua)
        ws.cell(row=row_idx, column=2, value=ua.sigla or '')
        ws.cell(row=row_idx, column=3, value=ua.municipio.nome if ua.municipio else '')
        ws.cell(row=row_idx, column=4, value=ua.municipio.codigo_ibge if ua.municipio else '')
        ws.cell(row=row_idx, column=5, value=ua.municipio.circunscricao if ua.municipio else '')
        ws.cell(row=row_idx, column=6, value=linha['total_pedidos'])
        ws.cell(row=row_idx, column=7, value=float(linha['total_qtd']))
        ws.cell(row=row_idx, column=8, value=float(linha['total_gasto']))

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="gastos_por_ua_sipat.xlsx"'
    wb.save(resp)
    return resp


def _nome_arquivo_gastos(request):
    """Sufixo do nome do arquivo a partir do filtro de data — pra exportar um arquivo
    por mês (ex.: gastos_sipat_2026-08.xlsx) e poder empilhar tudo depois numa pasta
    combinada no Power BI, sem um mês sobrescrever o anterior."""
    data_de = request.GET.get('data_de', '').strip()
    data_ate = request.GET.get('data_ate', '').strip()

    referencia = data_de or data_ate
    return referencia[:7] if referencia else 'completo'  # 'AAAA-MM-DD' -> 'AAAA-MM'


def _request_so_datas(request):
    """Isola só 'data_de'/'data_ate' da URL — evita que 'busca'/'grupo' de uma tela
    (item ou UA) vaze e filtre errado a aba da outra na exportação combinada."""
    qd = QueryDict(mutable=True)
    for chave in ('data_de', 'data_ate'):
        valor = request.GET.get(chave, '').strip()
        if valor:
            qd[chave] = valor
    return SimpleNamespace(GET=qd)


''' Exportação combinada (Gastos por Item + Gastos por UA num único XLSX, 2 abas) —
respeitando os filtros de data da URL, pra gerar um arquivo por mês e jogar numa
pasta que o Power BI combina automaticamente (Obter Dados → Pasta) '''
@login_required
def gastos_export_combinado_xlsx(request):
    req_datas = _request_so_datas(request)
    dados_item = _montar_gastos_por_item(req_datas)
    dados_ua = _montar_gastos_por_ua(req_datas)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E2D42', end_color='1E2D42', fill_type='solid')

    def cabecalho(ws, colunas):
        for col_idx, (titulo, largura) in enumerate(colunas, start=1):
            cel = ws.cell(row=1, column=col_idx, value=titulo)
            cel.font = header_font
            cel.fill = header_fill
            ws.column_dimensions[get_column_letter(col_idx)].width = largura
        ws.freeze_panes = 'A2'

    # ── aba 1: Gastos por Item ──
    ws1 = wb.active
    ws1.title = 'Gastos por Item'
    cabecalho(ws1, [
        ('E-Fisco', 16), ('Descrição', 60), ('Grupo', 22), ('Ano', 8),
        ('Mês', 14), ('Quantidade', 14), ('Preço Médio', 14), ('Gasto', 14),
    ])
    labels_grupo = dict(GrupoConsumo.choices)
    for row_idx, linha in enumerate(dados_item['linhas'], start=2):
        item = linha['item']
        ws1.cell(row=row_idx, column=1, value=item.efisco)
        ws1.cell(row=row_idx, column=2, value=item.descricao_efisco)
        ws1.cell(row=row_idx, column=3, value=labels_grupo.get(item.grupo_consumo, item.grupo_consumo))
        ws1.cell(row=row_idx, column=4, value=linha['ano'])
        ws1.cell(row=row_idx, column=5, value=MESES_PT[linha['mes']])
        ws1.cell(row=row_idx, column=6, value=float(linha['total_qtd']))
        ws1.cell(row=row_idx, column=7, value=float(item.preco_medio) if item.preco_medio is not None else None)
        ws1.cell(row=row_idx, column=8, value=float(linha['total_gasto']))

    # ── aba 2: Gastos por UA ──
    ws2 = wb.create_sheet('Gastos por UA')
    cabecalho(ws2, [
        ('UA', 46), ('Sigla', 14), ('Município', 28), ('Código IBGE', 14),
        ('Circunscrição', 26), ('Solicitações', 14), ('Quantidade', 14), ('Gasto', 14),
    ])
    for row_idx, linha in enumerate(dados_ua['linhas'], start=2):
        ua = linha['ua']
        ws2.cell(row=row_idx, column=1, value=ua.ua)
        ws2.cell(row=row_idx, column=2, value=ua.sigla or '')
        ws2.cell(row=row_idx, column=3, value=ua.municipio.nome if ua.municipio else '')
        ws2.cell(row=row_idx, column=4, value=ua.municipio.codigo_ibge if ua.municipio else '')
        ws2.cell(row=row_idx, column=5, value=ua.municipio.circunscricao if ua.municipio else '')
        ws2.cell(row=row_idx, column=6, value=linha['total_pedidos'])
        ws2.cell(row=row_idx, column=7, value=float(linha['total_qtd']))
        ws2.cell(row=row_idx, column=8, value=float(linha['total_gasto']))

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nome = f'gastos_sipat_{_nome_arquivo_gastos(request)}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{nome}"'
    wb.save(resp)
    return resp


''' Mural de Avisos '''
# Publicar um novo aviso no mural do DEMPAM (exibido em /home/)
@gerencia_dempam_required
def aviso_criar(request):
    if request.method == 'POST':
        form = AvisoForm(request.POST)
        if form.is_valid():
            aviso = form.save(commit=False)
            aviso.autor = request.user
            aviso.save()
            messages.success(request, 'Aviso publicado no mural com sucesso.')
            return redirect('dempam:homepage')
    else:
        form = AvisoForm()

    return render(request, 'dempam/aviso_form.html', {'form': form})


''' Editar um aviso existente do mural '''
@gerencia_dempam_required
def aviso_editar(request, pk):
    aviso = get_object_or_404(Aviso, pk=pk)
    if request.method == 'POST':
        form = AvisoForm(request.POST, instance=aviso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Aviso atualizado com sucesso.')
            return redirect('dempam:aviso_lista')
    else:
        form = AvisoForm(instance=aviso)

    return render(request, 'dempam/aviso_form.html', {'form': form, 'aviso': aviso, 'editando': True})


''' Listar avisos do mural (para gerenciar/excluir) '''
@gerencia_dempam_required
def aviso_lista(request):
    avisos = Aviso.objects.select_related('autor').all()
    return render(request, 'dempam/aviso_lista.html', {'avisos': avisos, 'now': timezone.now()})


''' Excluir um aviso do mural '''
@gerencia_dempam_required
@require_POST
def aviso_excluir(request, pk):
    aviso = get_object_or_404(Aviso, pk=pk)
    aviso.delete()
    messages.success(request, 'Aviso excluído com sucesso.')
    return redirect('dempam:aviso_lista')
