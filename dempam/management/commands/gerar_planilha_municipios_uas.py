"""
Gera uma planilha .xlsx pra preencher manualmente o município de cada UA que
ainda não tem um definido, com uma lista suspensa (validação de dados) com
os 185 municípios de Pernambuco.

Colunas da planilha:
  A  ID da UA        (não mexer — usado pra reimportar)
  B  UA
  C  Sigla
  D  Circunscrição/Prédio atual   (só referência)
  E  Município        (lista suspensa — preencher aqui)

Depois de preenchida, reimporte com:
    python manage.py importar_municipios_uas_xlsx "docs/UAs sem municipio.xlsx"

Uso:
    python manage.py gerar_planilha_municipios_uas
    python manage.py gerar_planilha_municipios_uas --saida "docs/outra.xlsx"
    python manage.py gerar_planilha_municipios_uas --todas   # inclui as que já têm município
"""
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from dempam.models import InfoUA, Municipio


class Command(BaseCommand):
    help = 'Gera planilha .xlsx com lista suspensa de municípios, pra atribuir UA -> município manualmente'

    def add_arguments(self, parser):
        parser.add_argument(
            '--saida', type=str, default=None,
            help='Caminho do arquivo de saída (padrão: docs/UAs sem município.xlsx)',
        )
        parser.add_argument(
            '--todas', action='store_true', default=False,
            help='Inclui também as UAs que já têm município definido (padrão: só as pendentes)',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise CommandError('openpyxl não instalado.')

        saida = Path(options['saida']) if options['saida'] else Path(settings.BASE_DIR) / 'docs' / 'UAs sem município.xlsx'

        qs = InfoUA.objects.select_related('municipio', 'circunscricao_predio').order_by('ua')
        if not options['todas']:
            qs = qs.filter(municipio__isnull=True)

        uas = list(qs)
        if not uas:
            self.stdout.write(self.style.WARNING('Nenhuma UA pra listar (todas já têm município? use --todas).'))
            return

        municipios = list(Municipio.objects.order_by('nome').values_list('nome', flat=True))

        wb = openpyxl.Workbook()

        # --- Aba principal ---
        ws = wb.active
        ws.title = 'UAs'
        cabecalho = ['ID (não mexer)', 'UA', 'Sigla', 'Circunscrição/Prédio atual', 'Município']
        ws.append(cabecalho)
        for col in range(1, len(cabecalho) + 1):
            cel = ws.cell(row=1, column=col)
            cel.font = Font(bold=True, color='FFFFFF')
            cel.fill = PatternFill('solid', fgColor='2563EB')

        for ua in uas:
            ws.append([
                ua.pk,
                ua.ua,
                ua.sigla,
                str(ua.circunscricao_predio) if ua.circunscricao_predio else '',
                ua.municipio.nome if ua.municipio else '',
            ])

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 30

        # --- Aba auxiliar com a lista de municípios (fonte da lista suspensa) ---
        ws_lista = wb.create_sheet('ListaMunicipios')
        for i, nome in enumerate(municipios, start=1):
            ws_lista.cell(row=i, column=1, value=nome)
        wb._sheets.remove(ws_lista)
        wb._sheets.insert(1, ws_lista)  # mantém referenciável, mesmo escondida
        ws_lista.sheet_state = 'hidden'

        dv = DataValidation(
            type='list',
            formula1=f"=ListaMunicipios!$A$1:$A${len(municipios)}",
            allow_blank=True,
            showDropDown=False,  # False = mostra a setinha (comportamento do Excel é invertido aqui)
        )
        dv.error = 'Escolha um município da lista.'
        dv.errorTitle = 'Município inválido'
        ws.add_data_validation(dv)
        dv.add(f'E2:E{len(uas) + 1}')

        saida.parent.mkdir(parents=True, exist_ok=True)
        wb.save(saida)

        self.stdout.write(self.style.SUCCESS(f'[OK] Planilha gerada: {saida}'))
        self.stdout.write(f'     {len(uas)} UA(s) listada(s), {len(municipios)} município(s) na lista suspensa.')
