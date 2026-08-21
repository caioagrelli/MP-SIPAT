"""
Reimporta a planilha gerada por `gerar_planilha_municipios_uas`, aplicando o
município escolhido (coluna E) em cada UA (identificada pela coluna A, ID).

A aba correta é localizada pelo cabeçalho (procura uma aba cuja primeira
linha tenha uma coluna "Município"), não pelo nome — a planilha preenchida
pelo usuário pode ter outras abas antes dela (ex.: "Página1", "Página2").

Nomes de município são comparados sem acento e sem diferenciar
maiúsculas/minúsculas. Linhas com a coluna E vazia são ignoradas (nada muda
pra essa UA). Nomes que não baterem nem por normalização nem pela tabela de
aliases são reportados como erro, sem aplicar nada errado.

Uso:
    python manage.py importar_municipios_uas_xlsx "docs/UAs sem município.xlsx"
"""
import re
import unicodedata
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from dempam.models import InfoUA, Municipio


def _normalizar(texto):
    sem_acento = unicodedata.normalize('NFKD', texto or '').encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'\s+', ' ', sem_acento).strip().lower()


# Nomes que apareceram na planilha preenchida e não batem letra-por-letra
# (por acento/grafia diferente) com o cadastro — mesmo padrão usado em
# aplicar_circunscricoes.py
_ALIASES = {
    'cabo santo agostinho': 'cabo de santo agostinho',
    'jaboatao dos grararapes': 'jaboatao dos guararapes',
    'itamaraca': 'ilha de itamaraca',
    'sao caetano': 'sao caitano',
    'camocim sao felix': 'camocim de sao felix',
    'santa maria boa vista': 'santa maria da boa vista',
    'santa maria cambuca': 'santa maria do cambuca',
    'belem de sao francisco': 'belem do sao francisco',
}


def _localizar_aba(wb):
    for aba in wb.sheetnames:
        ws = wb[aba]
        primeira_linha = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if 'Município' in primeira_linha:
            return ws
    return None


class Command(BaseCommand):
    help = 'Aplica o município escolhido na planilha (gerada por gerar_planilha_municipios_uas) em cada UA'

    def add_arguments(self, parser):
        parser.add_argument('arquivo', type=str, help='Caminho para o arquivo .xlsx preenchido')

    def handle(self, *args, **options):
        caminho = Path(options['arquivo'])
        if not caminho.exists():
            raise CommandError(f'Arquivo não encontrado: {caminho}')

        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl não instalado.')

        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = _localizar_aba(wb)
        if ws is None:
            raise CommandError('Nenhuma aba com coluna "Município" encontrada na planilha.')

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        municipios_por_norm = {_normalizar(m.nome): m for m in Municipio.objects.all()}

        atualizadas = 0
        sem_escolha = 0
        ids_invalidos = []
        municipios_invalidos = []

        for row in rows:
            if not row or row[0] is None:
                continue
            ua_id = row[0]
            municipio_nome = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else '')

            if not municipio_nome:
                sem_escolha += 1
                continue

            chave = _normalizar(municipio_nome)
            chave = _ALIASES.get(chave, chave)
            municipio = municipios_por_norm.get(chave)
            if not municipio:
                municipios_invalidos.append((ua_id, municipio_nome))
                continue

            ua = InfoUA.objects.filter(pk=ua_id).first()
            if not ua:
                ids_invalidos.append(ua_id)
                continue

            if ua.municipio_id != municipio.pk:
                ua.municipio = municipio
                ua.save(update_fields=['municipio'])
            atualizadas += 1

        self.stdout.write(self.style.SUCCESS(f'[OK] UAs atualizadas: {atualizadas}'))
        self.stdout.write(self.style.WARNING(f'[--] Linhas sem município escolhido: {sem_escolha}'))

        if municipios_invalidos:
            self.stdout.write(self.style.ERROR(f'[XX] Município não encontrado no cadastro: {len(municipios_invalidos)}'))
            for ua_id, nome in municipios_invalidos[:30]:
                self.stdout.write(f'  UA #{ua_id}: {nome!r}')

        if ids_invalidos:
            self.stdout.write(self.style.ERROR(f'[XX] ID de UA não encontrado: {len(ids_invalidos)}'))
            for ua_id in ids_invalidos[:30]:
                self.stdout.write(f'  {ua_id}')
